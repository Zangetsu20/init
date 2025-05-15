from flask import Flask, render_template, request, redirect, flash, url_for, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_bcrypt import Bcrypt
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from markupsafe import escape #XSS защита
from sqlalchemy import UniqueConstraint
# Добавляем в начало файла (после других импортов)
from openpyxl import Workbook
from io import BytesIO
from flask import send_file
from datetime import datetime
from flask import session
import zipfile
import tempfile
import os
#from flask_wtf.csrf import CSRFProtect #CSRF защита
#from flask_wtf.csrf import generate_csrf

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///newf.db'
import os
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', os.urandom(24).hex())
#csrf = CSRFProtect(app) #csrf защита
db = SQLAlchemy(app)
bcrypt = Bcrypt(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'

#@app.context_processor
#def inject_csrf():
#    return dict(csrf_token=generate_csrf)  

# User model
class User(db.Model, UserMixin):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(150), unique=True, nullable=False)
    password = db.Column(db.String(60), nullable=False)
    teacher_id = db.Column(db.Integer, db.ForeignKey('teachers.id')) 

# Graf model
#class Graf(db.Model):
#    id = db.Column(db.Integer, primary_key=True)
#   firstname = db.Column(db.String(300), nullable=False)
#    lastname = db.Column(db.String(300), nullable=False)
#    d = db.Column(db.String(300), nullable=False)

# создание teachers таблицы
class Teachers(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    first_name = db.Column(db.Text)
    last_name = db.Column(db.Text)
    
# создание teachers_subject таблицы
class Teacher_subject(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    teacher_id = db.Column(db.Integer, nullable=False)
    subject_id = db.Column(db.Integer, nullable=False)

# создание subject таблицы
class Subjects(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    subject_name = db.Column(db.Text, nullable=False, unique=True)

# создание teachers_subject таблицы
class Teacher_subject_class(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    teacher_subject_id = db.Column(db.Integer, nullable=False)
    class_id = db.Column(db.Integer, nullable=False)    

# создание classes таблицы
class Classes(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    class_name = db.Column(db.Text, nullable=False, unique=True)

# создание students таблицы
class Students(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    first_name = db.Column(db.Text, nullable=False)
    last_name = db.Column(db.Text, nullable=False)   
    class_id = db.Column(db.Integer)
    
# создание grades таблицы
class Grades(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, nullable=False)
    teacher_subject_class_id = db.Column(db.Integer, nullable=False)   
    grade = db.Column(db.Text, nullable=True)
    date = db.Column(db.Date, nullable=False)
    __table_args__ = (UniqueConstraint('student_id', 'date', 'teacher_subject_class_id', name='_student_date_tsc_uc'),)


# Create database tables (применить создание таблиц, должно быть в конце после создания всех таблиц)
with app.app_context():
    db.create_all()

# User loader for Flask-Login
@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


#тест выода\редактирования
#@csrf.exempt #разкоментировать для отключения csrf 
@app.route('/update_teacher', methods=['POST'])
def update_teacher():
    try:
        data = request.get_json()
        
        # Проверка наличия всех необходимых полей
        if not data or 'id' not in data or 'field' not in data or 'value' not in data:
            return jsonify(error="Неверный формат данных: отсутствуют обязательные поля"), 400
        
        teacher = Teachers.query.get(data['id'])
        if not teacher:
            return jsonify(error="Учитель не найден"), 404
          
        # Обновляем поле 
        if data['field'] == 'first_name':
            teacher.first_name = escape(data['value'])
        elif data['field'] == 'last_name':
            teacher.last_name = escape(data['value']) 
        else:
            return jsonify(error="Недопустимое поле"), 400
        
        db.session.commit()
        return jsonify(success=True)
    
    except Exception as e:
        db.session.rollback()
        return jsonify(error=str(e)), 500

@app.route('/get_class_data')
@login_required
def get_class_data():
    try:
        class_id = request.args.get('class_id', type=int)
        subject_id = request.args.get('subject_id', type=int)
        
        if not class_id or not subject_id:
            return jsonify({'error': 'Не указаны ID класса или предмета'}), 400
        
        # Для администратора
        if current_user.teacher_id == 12:
            teacher_subject = Teacher_subject.query.filter_by(
                subject_id=subject_id
            ).first()
        else:
            # Для обычного учителя
            teacher_subject = Teacher_subject.query.filter_by(
                teacher_id=current_user.teacher_id,
                subject_id=subject_id
            ).first()
        
        if not teacher_subject:
            return jsonify({'error': 'Преподаватель не ведет этот предмет'}), 404
        
        tsc = Teacher_subject_class.query.filter_by(
            teacher_subject_id=teacher_subject.id,
            class_id=class_id
        ).first()
        
        if not tsc:
            return jsonify({'error': 'Не найдена связь преподаватель-класс'}), 404
        
        # Получаем первого ученика в классе
        student = Students.query.filter_by(class_id=class_id).first()
        
        return jsonify({
            'teacher_subject_class_id': tsc.id,
            'first_student_id': student.id if student else None
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route("/get_teacher_id")
@login_required
def get_teacher_id():
    subject_id = request.args.get("subject_id", type=int)  # Получаем subject_id из запроса
    teacher_id = current_user.teacher_id  # teacher_id из авторизованного пользователя

    if not teacher_id:
        return jsonify({
            "success": False,
            "error": "Пользователь не привязан к преподавателю"
        }), 400

    if not subject_id:
        return jsonify({
            "success": False,
            "error": "Не указан subject_id"
        }), 400

    # Проверяем, что преподаватель ведёт данный предмет
    teacher_subject = Teacher_subject.query.filter_by(
        teacher_id=teacher_id,
        subject_id=subject_id
    ).first()

    if not teacher_subject:
        return jsonify({
            "success": False,
            "error": "Преподаватель не ведёт данный предмет"
        }), 404

    return jsonify({
        "success": True,
        "teacher_id": teacher_id,
        "subject_id": subject_id
    })

@app.route('/update_grade', methods=['POST'])
@login_required
def update_grade():
    try:
        data = request.get_json()
        
        # Проверяем обязательные поля
        required_fields = ['student_id', 'date', 'teacher_subject_class_id', 'new_grade']
        if not all(field in data for field in required_fields):
            return jsonify({'success': False, 'error': 'Не хватает обязательных полей'}), 400
        
        # Преобразуем дату из формата DD.MM в объект date
        date_str = data['date']
        try:
            date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError as e:
            return jsonify({'success': False, 'error': f'Неверный формат даты: {str(e)}'}), 400
        
        # Если student_id = 0, просто создаем запись в базе без привязки к ученику
        if data['student_id'] == 0:
            # Находим любого ученика в этом классе
            tsc = Teacher_subject_class.query.get(data['teacher_subject_class_id'])
            if not tsc:
                return jsonify({'success': False, 'error': 'Класс не найден'}), 404
                
            student = Students.query.filter_by(class_id=tsc.class_id).first()
            if not student:
                return jsonify({'success': False, 'error': 'В классе нет учеников'}), 404
                
            data['student_id'] = student.id
        
        # Ищем существующую оценку
        grade = Grades.query.filter_by(
            student_id=data['student_id'],
            teacher_subject_class_id=data['teacher_subject_class_id'],
            date=date_obj
        ).first()
        
        if grade:
            # Обновляем существующую оценку
            grade.grade = data['new_grade']
        else:
            # Создаем новую запись
            grade = Grades(
                student_id=data['student_id'],
                teacher_subject_class_id=data['teacher_subject_class_id'],
                grade=data['new_grade'],
                date=date_obj
            )
            db.session.add(grade)
        
        db.session.commit()
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/get_first_student')
@login_required
def get_first_student():
    try:
        class_id = request.args.get('class_id', type=int)
        if not class_id:
            return jsonify({'error': 'Не указан ID класса'}), 400
        
        # Находим первого ученика в классе
        student = Students.query.filter_by(class_id=class_id).first()
        if not student:
            return jsonify({'error': 'В классе нет учеников'}), 404
        
        # Находим teacher_subject_class_id для текущего учителя и класса
        if current_user.teacher_id == 12:  # Админ
            tsc = Teacher_subject_class.query.filter_by(class_id=class_id).first()
        else:
            # Для обычного учителя
            teacher_subject = Teacher_subject.query.filter_by(
                teacher_id=current_user.teacher_id
            ).first()
            if not teacher_subject:
                return jsonify({'error': 'Учитель не ведет предметы'}), 404
            
            tsc = Teacher_subject_class.query.filter_by(
                teacher_subject_id=teacher_subject.id,
                class_id=class_id
            ).first()
        
        if not tsc:
            return jsonify({'error': 'Не найдена связь учитель-предмет-класс'}), 404
        
        return jsonify({
            'student_id': student.id,
            'teacher_subject_class_id': tsc.id
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

#что это?
@app.route("/index")
@app.route("/")
def index():
    return render_template('index.html')

#
#данные к тестовой первой таблицы из БД
@app.route("/posts")
@login_required
def posts():
    # Очищаем session storage при первом заходе
    if not request.args.get('year') and not request.args.get('month'):
        session.pop('activeSubjectTab', None)
        session.pop('activeClassTab', None)
    
    # Получаем параметры фильтрации
    selected_year = request.args.get('year', type=int)
    selected_month = request.args.get('month', type=int)
    
        # Если параметров нет, проверяем localStorage через cookies
    if selected_year is None or selected_month is None:
        stored_year = request.cookies.get('selectedYear')
        stored_month = request.cookies.get('selectedMonth')
        if stored_year and stored_month:
            selected_year = int(stored_year)
            selected_month = int(stored_month)
    # Если все еще нет - используем текущую дату
    if selected_year is None:
        selected_year = datetime.now().year
    if selected_month is None:
        selected_month = datetime.now().month
        
    # Проверяем, является ли пользователь администратором
    is_admin = current_user.teacher_id == 12
    
    if is_admin:
        subjects = Subjects.query.order_by(Subjects.subject_name).all()
    else:
        if not current_user.teacher_id:
            flash('У вас нет привязанного профиля учителя', 'error')
            return redirect(url_for('index'))
        
        teacher_id = current_user.teacher_id
        teacher_subjects = Teacher_subject.query.filter_by(teacher_id=teacher_id).all()
        subject_ids = [ts.subject_id for ts in teacher_subjects]
        subjects = Subjects.query.filter(Subjects.id.in_(subject_ids)).order_by(Subjects.subject_name).all()
    
    # Получаем доступные годы из базы
    min_year = db.session.query(db.extract('year', db.func.min(Grades.date))).scalar() or datetime.now().year
    max_year = db.session.query(db.extract('year', db.func.max(Grades.date))).scalar() or datetime.now().year
    available_years = list(range(min_year, max_year + 1))
    
    subjects_data = []
    
    for subject in subjects:
        if is_admin:
            teacher_subjects = Teacher_subject.query.filter_by(subject_id=subject.id).all()
            teacher_subject_ids = [ts.id for ts in teacher_subjects]
            class_links = Teacher_subject_class.query.filter(
                Teacher_subject_class.teacher_subject_id.in_(teacher_subject_ids)
            ).all()
        else:
            teacher_subject = Teacher_subject.query.filter_by(
                teacher_id=current_user.teacher_id,
                subject_id=subject.id
            ).first()
            if not teacher_subject:
                continue
            
            class_links = Teacher_subject_class.query.filter_by(
                teacher_subject_id=teacher_subject.id
            ).all()
        
        classes_data = []
        for cl in class_links:
            class_info = Classes.query.get(cl.class_id)
            if not class_info:
                continue
            
            # Получаем имя преподавателя (для админа)
            teacher_name = ""
            if is_admin:
                ts = Teacher_subject.query.get(cl.teacher_subject_id)
                if ts:
                    teacher = Teachers.query.get(ts.teacher_id)
                    if teacher:
                        teacher_name = f"{teacher.last_name} {teacher.first_name}"
            
            # Получаем студентов
            students = Students.query.filter_by(class_id=class_info.id)\
                .order_by(Students.last_name, Students.first_name).all()
            
            students_grades = []
            all_dates = set()
            
            for student in students:
                # Фильтруем оценки по выбранному году и месяцу
                grades = Grades.query.filter(
                    Grades.student_id == student.id,
                    Grades.teacher_subject_class_id == cl.id,
                    db.extract('year', Grades.date) == selected_year,
                    db.extract('month', Grades.date) == selected_month
                ).order_by(Grades.date).all()
                
                grades_dict = {}
                for grade in grades:
                    day_str = grade.date.strftime('%d')  # Только день
                    grades_dict[day_str] = grade.grade
                    all_dates.add(day_str)
                
                students_grades.append({
                    'last_name': student.last_name,
                    'first_name': student.first_name,
                    'student_id': student.id,
                    'teacher_subject_class_id': cl.id,
                    'grades': grades_dict
                })
            
            sorted_dates = sorted(all_dates)
            
            classes_data.append({
                'class_id': class_info.id,
                'class_name': class_info.class_name,
                'teacher_name': teacher_name,
                'students': students_grades,
                'dates': sorted_dates
            })
        
        subjects_data.append({
            'subject_id': subject.id,
            'subject_name': subject.subject_name,
            'classes': classes_data  # Может быть пустым
})
    
    return render_template('posts.html', 
                         subjects_data=subjects_data,
                         is_admin=is_admin,
                         available_years=available_years,
                         available_months=list(range(1, 13)),
                         selected_year=selected_year,
                         selected_month=selected_month)

@app.template_filter('month_name')
def month_name_filter(month_num):
    months = [
        'Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
        'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь'
    ]
    return months[month_num - 1] if 1 <= month_num <= 12 else str(month_num)
    
def get_available_years():
    # Получаем минимальный и максимальный год из базы
    min_year = db.session.query(db.extract('year', db.func.min(Grades.date))).scalar()
    max_year = db.session.query(db.extract('year', db.func.max(Grades.date))).scalar()
    return list(range(min_year, max_year + 1)) if min_year and max_year else [datetime.now().year]

@app.route("/create")
@login_required
def create():
    # Очищаем session storage при первом заходе
    if not request.args.get('year') and not request.args.get('month'):
        session.pop('activeSubjectTab', None)
        session.pop('activeClassTab', None)
    
    # Получаем параметры фильтрации
    selected_year = request.args.get('year', type=int)
    selected_month = request.args.get('month', type=int)
    
    # Если параметров нет, проверяем localStorage через cookies
    if selected_year is None or selected_month is None:
        stored_year = request.cookies.get('selectedYear')
        stored_month = request.cookies.get('selectedMonth')
        if stored_year and stored_month:
            selected_year = int(stored_year)
            selected_month = int(stored_month)
    # Если все еще нет - используем текущую дату
    if selected_year is None:
        selected_year = datetime.now().year
    if selected_month is None:
        selected_month = datetime.now().month
        
    # Проверяем, является ли пользователь администратором
    is_admin = current_user.teacher_id == 12
    
    if is_admin:
        subjects = Subjects.query.order_by(Subjects.subject_name).all()
    else:
        if not current_user.teacher_id:
            flash('У вас нет привязанного профиля учителя', 'error')
            return redirect(url_for('index'))
        
        teacher_id = current_user.teacher_id
        teacher_subjects = Teacher_subject.query.filter_by(teacher_id=teacher_id).all()
        subject_ids = [ts.subject_id for ts in teacher_subjects]
        subjects = Subjects.query.filter(Subjects.id.in_(subject_ids)).order_by(Subjects.subject_name).all()
    
    # Получаем доступные годы из базы
    min_year = db.session.query(db.extract('year', db.func.min(Grades.date))).scalar() or datetime.now().year
    max_year = db.session.query(db.extract('year', db.func.max(Grades.date))).scalar() or datetime.now().year
    available_years = list(range(min_year, max_year + 1))
    
    subjects_data = []
    
    for subject in subjects:
        if is_admin:
            teacher_subjects = Teacher_subject.query.filter_by(subject_id=subject.id).all()
            teacher_subject_ids = [ts.id for ts in teacher_subjects]
            class_links = Teacher_subject_class.query.filter(
                Teacher_subject_class.teacher_subject_id.in_(teacher_subject_ids)
            ).all()
        else:
            teacher_subject = Teacher_subject.query.filter_by(
                teacher_id=current_user.teacher_id,
                subject_id=subject.id
            ).first()
            if not teacher_subject:
                continue
            
            class_links = Teacher_subject_class.query.filter_by(
                teacher_subject_id=teacher_subject.id
            ).all()
        
        classes_data = []
        for cl in class_links:
            class_info = Classes.query.get(cl.class_id)
            if not class_info:
                continue
            
            # Получаем имя преподавателя (для админа)
            teacher_name = ""
            if is_admin:
                ts = Teacher_subject.query.get(cl.teacher_subject_id)
                if ts:
                    teacher = Teachers.query.get(ts.teacher_id)
                    if teacher:
                        teacher_name = f"{teacher.last_name} {teacher.first_name}"

            # Получаем всех студентов класса
            students = Students.query.filter_by(class_id=class_info.id)\
                .order_by(Students.last_name, Students.first_name).all()
            
            if not students:
                classes_data.append({
                    'class_id': class_info.id,
                    'class_name': class_info.class_name,
                    'teacher_name': teacher_name,
                    'students': [],
                    'max_day': 0
                })
                continue
            
            # Получаем все оценки для этого класса и предмета за выбранный период
            all_grades = Grades.query.filter(
                Grades.teacher_subject_class_id == cl.id,
                db.extract('year', Grades.date) == selected_year,
                db.extract('month', Grades.date) == selected_month
            ).all()
            
            # Находим максимальный день в классе (по всем студентам)
            max_day = max([grade.date.day for grade in all_grades]) if all_grades else 0
            
            students_data = []
            for student in students:
                # Фильтруем оценки только для текущего студента
                student_grades = [g for g in all_grades if g.student_id == student.id]
                
                # Считаем пропуски ("н")
                absent_count = sum(1 for g in student_grades if g.grade == 'н')
                # Рассчитываем процент посещаемости
                attendance_percent = 0
                if max_day > 0:
                    attendance_percent = round(((max_day - absent_count) / max_day) * 100)
                
                students_data.append({
                    'last_name': student.last_name,
                    'first_name': student.first_name,
                    'student_id': student.id,
                    'absent_count': absent_count,
                    'has_records': len(student_grades) > 0,
                    'attendance_percent': attendance_percent  # Добавляем расчетный процент
                })
            
            classes_data.append({
                'class_id': class_info.id,
                'class_name': class_info.class_name,
                'teacher_name': teacher_name,
                'students': students_data,
                'max_day': max_day
            })
        
        if classes_data:
            subjects_data.append({
                'subject_id': subject.id,
                'subject_name': subject.subject_name,
                'classes': classes_data
            })
    
    return render_template('create.html', 
                         subjects_data=subjects_data,
                         is_admin=is_admin,
                         available_years=available_years,
                         available_months=list(range(1, 13)),
                         selected_year=selected_year,
                         selected_month=selected_month)

@app.route("/about")
def about():
    return render_template('about.html')

@app.route("/abot")
@login_required
def abot():
    # Очищаем session storage при первом заходе
    if not request.args.get('year') and not request.args.get('month'):
        session.pop('activeSubjectTab', None)
        session.pop('activeClassTab', None)
    
    # Получаем параметры фильтрации
    selected_year = request.args.get('year', type=int)
    selected_month = request.args.get('month', type=int)
    
    # Если параметров нет, проверяем localStorage через cookies
    if selected_year is None or selected_month is None:
        stored_year = request.cookies.get('selectedYear')
        stored_month = request.cookies.get('selectedMonth')
        if stored_year and stored_month:
            selected_year = int(stored_year)
            selected_month = int(stored_month)
    # Если все еще нет - используем текущую дату
    if selected_year is None:
        selected_year = datetime.now().year
    if selected_month is None:
        selected_month = datetime.now().month
        
    # Проверяем, является ли пользователь администратором
    is_admin = current_user.teacher_id == 12
    
    if is_admin:
        subjects = Subjects.query.order_by(Subjects.subject_name).all()
    else:
        if not current_user.teacher_id:
            flash('У вас нет привязанного профиля учителя', 'error')
            return redirect(url_for('index'))
        
        teacher_id = current_user.teacher_id
        teacher_subjects = Teacher_subject.query.filter_by(teacher_id=teacher_id).all()
        subject_ids = [ts.subject_id for ts in teacher_subjects]
        subjects = Subjects.query.filter(Subjects.id.in_(subject_ids)).order_by(Subjects.subject_name).all()
    
    # Получаем доступные годы из базы
    min_year = db.session.query(db.extract('year', db.func.min(Grades.date))).scalar() or datetime.now().year
    max_year = db.session.query(db.extract('year', db.func.max(Grades.date))).scalar() or datetime.now().year
    available_years = list(range(min_year, max_year + 1))
    
    subjects_data = []
    
    for subject in subjects:
        if is_admin:
            teacher_subjects = Teacher_subject.query.filter_by(subject_id=subject.id).all()
            teacher_subject_ids = [ts.id for ts in teacher_subjects]
            class_links = Teacher_subject_class.query.filter(
                Teacher_subject_class.teacher_subject_id.in_(teacher_subject_ids)
            ).all()
        else:
            teacher_subject = Teacher_subject.query.filter_by(
                teacher_id=current_user.teacher_id,
                subject_id=subject.id
            ).first()
            if not teacher_subject:
                continue
            
            class_links = Teacher_subject_class.query.filter_by(
                teacher_subject_id=teacher_subject.id
            ).all()
        
        classes_data = []
        for cl in class_links:
            class_info = Classes.query.get(cl.class_id)
            if not class_info:
                continue
            
            # Получаем имя преподавателя (для админа)
            teacher_name = ""
            if is_admin:
                ts = Teacher_subject.query.get(cl.teacher_subject_id)
                if ts:
                    teacher = Teachers.query.get(ts.teacher_id)
                    if teacher:
                        teacher_name = f"{teacher.last_name} {teacher.first_name}"

            # Получаем всех студентов класса
            students = Students.query.filter_by(class_id=class_info.id)\
                .order_by(Students.last_name, Students.first_name).all()
            
            if not students:
                classes_data.append({
                    'class_id': class_info.id,
                    'class_name': class_info.class_name,
                    'teacher_name': teacher_name,
                    'students': [],
                    'max_day': 0
                })
                continue
            
            # Получаем все оценки для этого класса и предмета за выбранный период
            all_grades = Grades.query.filter(
                Grades.teacher_subject_class_id == cl.id,
                db.extract('year', Grades.date) == selected_year,
                db.extract('month', Grades.date) == selected_month
            ).all()
            
            # Находим максимальный день в классе (по всем студентам)
            max_day = max([grade.date.day for grade in all_grades]) if all_grades else 0
            
            students_data = []
            for student in students:
                # Фильтруем оценки только для текущего студента
                student_grades = [g for g in all_grades if g.student_id == student.id]
                
                # Считаем пропуски ("н")
                absent_count = sum(1 for g in student_grades if g.grade == 'н')
                
                # Собираем только числовые оценки
                numeric_grades = [
                    int(g.grade) for g in student_grades 
                    if g.grade.strip() in {'1', '2', '3', '4', '5', '6', '7', '8', '9','10', '11','12'}
                ]
                
                empty_grades = [g.grade for g in student_grades if (g.grade or '').strip() == '']
                
                # Количество дней с оценками (только пробелы)
                graded_spaces = len(empty_grades)
                # Количество дней с оценками (только цифры)
                graded_days = len(numeric_grades)

                # Среднее грубо: сумма оценок / количество оцененных дней
                avg_grade_rude = None
                if graded_days > 0:
                    avg_grade_rude = round(sum(numeric_grades) / graded_days, 2)
                
                total_days = max_day  # Всего дней
                present_days = total_days - absent_count if total_days else 0
                    
                avg_grade_all = None
                avg_grade_present = None
                activity = 0
                    
                if numeric_grades:
                    total_grade_sum = sum(numeric_grades)
                        
                    if total_days > 0:
                            avg_grade_all = round(total_grade_sum / total_days, 2)
                            
                    if present_days > 0:
                            avg_grade_present = round(total_grade_sum / present_days, 2)
                            
                    if present_days > 0:
                            activity = round(graded_days/present_days,2)*100
                    else:
                            activity = 0
                    if student.id ==166:
                        print('Янг',graded_days,present_days, graded_spaces,activity)
                # Рассчитываем процент посещаемости
                attendance_percent = 0
                if max_day > 0:
                    attendance_percent = round(((max_day - absent_count) / max_day) * 100)
                    
                    
                
                students_data.append({
                    'last_name': student.last_name,
                    'first_name': student.first_name,
                    'student_id': student.id,
                    'absent_count': absent_count,
                    'has_records': len(student_grades) > 0,
                    'attendance_percent': attendance_percent,  # Добавляем расчетный процент
                    'avg_grade_all': avg_grade_all,        # Среднее по всем дням
                    'avg_grade_present': avg_grade_present, # Среднее без учёта "н"
                    'avg_grade_rude': avg_grade_rude, #среднее только по оцененным дням
                    'activity': activity #активнотсь
                })
            
            classes_data.append({
                'class_id': class_info.id,
                'class_name': class_info.class_name,
                'teacher_name': teacher_name,
                'students': students_data,
                'max_day': max_day
            })
        
        if classes_data:
            subjects_data.append({
                'subject_id': subject.id,
                'subject_name': subject.subject_name,
                'classes': classes_data
            })
    
    return render_template('abot.html', 
                         subjects_data=subjects_data,
                         is_admin=is_admin,
                         available_years=available_years,
                         available_months=list(range(1, 13)),
                         selected_year=selected_year,
                         selected_month=selected_month)

@app.route("/register", methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        if User.query.filter_by(username=username).first():
            flash('Имя пользователя уже занято!', 'error')
            return render_template('register.html')
        
        hashed_password = bcrypt.generate_password_hash(password).decode('utf-8')
        user = User(username=username, password=hashed_password)
        
        try:
            db.session.add(user)
            db.session.commit()
            flash('Регистрация прошла успешно! Пожалуйста, войдите.', 'success')
            return redirect(url_for('login'))
        except:
            flash('При регистрации произошла ошибка!', 'error')
            return render_template('register.html')
    
    return render_template('register.html')

@app.route("/login", methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username).first()
        
        if user and bcrypt.check_password_hash(user.password, password):
            login_user(user)
            flash('Вы успешно вошли!', 'success')
            return redirect(url_for('index'))
        else:
            flash('Неверное имя пользователя или пароль!', 'error')
            return render_template('login.html')
    
    return render_template('login.html')

@app.route("/logout")
@login_required
def logout():
    logout_user()
    flash('Вы вышли из системы!', 'success')
    return redirect(url_for('index'))


@app.route('/download_all_zip')
@login_required
def download_all_zip():
    try:
        # Создаем временный файл в памяти
        buffer = BytesIO()
        
        # Получаем доступные предметы
        if current_user.teacher_id == 12:  # Админ
            subjects = Subjects.query.all()
        else:
            teacher_subjects = Teacher_subject.query.filter_by(
                teacher_id=current_user.teacher_id
            ).all()
            subject_ids = [ts.subject_id for ts in teacher_subjects]
            subjects = Subjects.query.filter(Subjects.id.in_(subject_ids)).all()

        # Создаем ZIP в памяти
        with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for subject in subjects:
                # Создаем Excel файл в памяти
                excel_buffer = BytesIO()
                wb = Workbook()
                if 'Sheet' in wb.sheetnames:
                    wb.remove(wb['Sheet'])

                # Получаем классы для предмета
                if current_user.teacher_id == 12:
                    teacher_subjects = Teacher_subject.query.filter_by(subject_id=subject.id).all()
                    teacher_subject_ids = [ts.id for ts in teacher_subjects]
                    class_links = Teacher_subject_class.query.filter(
                        Teacher_subject_class.teacher_subject_id.in_(teacher_subject_ids)
                    ).all()
                else:
                    teacher_subject = Teacher_subject.query.filter_by(
                        teacher_id=current_user.teacher_id,
                        subject_id=subject.id
                    ).first()
                    if not teacher_subject:
                        continue
                    class_links = Teacher_subject_class.query.filter_by(
                        teacher_subject_id=teacher_subject.id
                    ).all()

                # Заполняем Excel данными
                for cl in class_links:
                    class_info = Classes.query.get(cl.class_id)
                    if not class_info:
                        continue
                    
                    ws = wb.create_sheet(title=class_info.class_name[:31])
                    ws.append(["Предмет:", subject.subject_name])
                    ws.append(["Класс:", class_info.class_name])
                    ws.append([])
                    
                    students = Students.query.filter_by(class_id=class_info.id)\
                        .order_by(Students.last_name, Students.first_name).all()
                    
                    grades = Grades.query.filter_by(teacher_subject_class_id=cl.id)\
                        .order_by(Grades.date).all()
                    dates = sorted({grade.date.strftime('%d.%m.%Y') for grade in grades})
                    
                    ws.append(["№", "Фамилия", "Имя"] + dates)
                    
                    for i, student in enumerate(students, 1):
                        row = [i, student.last_name, student.first_name]
                        grades_dict = {
                            g.date.strftime('%d.%m.%Y'): g.grade 
                            for g in Grades.query.filter_by(
                                student_id=student.id,
                                teacher_subject_class_id=cl.id
                            ).all()
                        }
                        row.extend(grades_dict.get(date, "") for date in dates)
                        ws.append(row)
                    
                    ws.append([])

                # Сохраняем Excel в буфер
                if len(wb.sheetnames) > 0:
                    wb.save(excel_buffer)
                    excel_buffer.seek(0)
                    zipf.writestr(f"{subject.subject_name}.xlsx", excel_buffer.getvalue())

        buffer.seek(0)
        
        # Отправляем ZIP-архив
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"grades_{datetime.now().strftime('%Y%m%d_%H%M')}.zip",
            mimetype='application/zip'
        )

    except Exception as e:
        app.logger.error(f"Error generating ZIP: {str(e)}")
        flash(f'Ошибка при создании архива: {str(e)}', 'error')
        return redirect(url_for('posts'))

@app.route("/add_student", methods=["POST"])
@login_required
def add_student():
    firstname = request.form.get("firstname")
    lastname = request.form.get("lastname")
    classid = request.form.get("classid")
    print(classid,lastname,firstname)

    if not firstname or not lastname:
        flash("Необходимо заполнить все поля", "error")
        return redirect(url_for('posts'))

    try:
        new_student = Students(
            first_name=escape(firstname),
            last_name=escape(lastname),
            class_id=escape(classid)
        )
        db.session.add(new_student)
        db.session.commit()
        flash("Студент успешно добавлен", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Ошибка при добавлении студента: {str(e)}", "error")

    return redirect(url_for('posts'))

@app.route("/add_class", methods=["POST"])
@login_required
def add_class():
    classname = request.form.get("classname")
    subject_id = request.form.get("subject_id")
    teacher_id = request.form.get("teacher_id")
    if not all([subject_id, teacher_id]):
        flash("Невозможно опознать учителя или предмет", "error")
        return redirect(url_for('posts'))

    if not classname:
        flash("Необходимо заполнить все поля", "error")
        return redirect(url_for('posts'))
    

    teacher_subject = Teacher_subject.query.filter_by(
    teacher_id=teacher_id,
    subject_id=subject_id
        ).first()
    if teacher_subject:
            teacher_subject_id = teacher_subject.id
            
    try:
        new_class = Classes(
            class_name=escape(classname)
        )
        db.session.add(new_class)                 
        db.session.commit()            
        new_teacher_subject_class = Teacher_subject_class(
        teacher_subject_id = teacher_subject.id,
        class_id = new_class.id
        )
        db.session.add(new_teacher_subject_class)
        db.session.commit()

        
        print("ID новой связки:", new_teacher_subject_class.id)
        print("ID нового класса:", new_class.id)
        flash("Класс успешно добавлен", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Ошибка при добавлении класса: {str(e)}", "error")

    return redirect(url_for('posts'))

@app.route("/add_subject", methods=["POST"])
@login_required
def add_subject():
    subjectsname = request.form.get("subjectsname")
    teacher_id = request.form.get("teacher_id")
    print(subjectsname,teacher_id)
    if not subjectsname:
        flash("Необходимо заполнить все поля", "error")
        return redirect(url_for('posts'))

    try:
        # Создаём новый предмет
        new_subject = Subjects(subject_name=escape(subjectsname))
        db.session.add(new_subject)
        db.session.flush()  # Чтобы получить ID нового предмета

        # Создаём связь между преподавателем и предметом
        new_teacher_subject = Teacher_subject(
            teacher_id=teacher_id,
            subject_id=new_subject.id
        )
        db.session.add(new_teacher_subject)

        db.session.commit()
        flash("Предмет успешно добавлен!", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Ошибка при добавлении предмета: {str(e)}", "error")

    return redirect(url_for('posts'))

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
    
#if __name__ == '__main__':
#    app.run(debug=True)